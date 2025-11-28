#!/usr/bin/env python3
"""
Generate the Excel template for Ops Manager alert configurations.
Run this script once to create the default configuration file.
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

# Alert configurations for Ops Manager
# Note: Some Atlas alerts (CPS_* for Cloud Backup) are not available in Ops Manager
ALERT_CONFIGS = [
    # Replication Alerts
    {
        "name": "Oplog Window",
        "category": "Replica Set",
        "low_threshold": "< 24h for 5 minutes",
        "high_threshold": "< 12h for 5 minutes",
        "description": "Oplog window is running out. Risk of secondary falling too far behind.",
    },
    {
        "name": "Replication Lag",
        "category": "Replica Set",
        "low_threshold": "> 60s for 5 minutes",
        "high_threshold": "> 120s for 5 minutes",
        "description": "Secondary is lagging behind the primary.",
    },
    {
        "name": "Replica set has no primary",
        "category": "Replica Set",
        "low_threshold": "5 minutes",
        "high_threshold": "1 minute",
        "description": "No primary detected in replica set. Write operations will fail.",
    },
    {
        "name": "Replica set elected a new primary",
        "category": "Replica Set",
        "low_threshold": "Any occurrence",
        "high_threshold": None,
        "description": "A new primary was elected. May indicate failover event.",
    },

    # Host Alerts
    {
        "name": "Host is Down",
        "category": "Host",
        "low_threshold": "5 minutes",
        "high_threshold": "1 minute",
        "description": "Host is unreachable or not responding.",
    },
    {
        "name": "Host is recovering",
        "category": "Host",
        "low_threshold": "Any occurrence",
        "high_threshold": None,
        "description": "Host entered recovering state. Cannot serve reads or writes.",
    },

    # Disk I/O Alerts
    {
        "name": "Disk read IOPS on Data Partition",
        "category": "Host",
        "low_threshold": "> 4000 for 5 minutes",
        "high_threshold": "> 8000 for 5 minutes",
        "description": "High disk read operations per second.",
    },
    {
        "name": "Disk write IOPS on Data Partition",
        "category": "Host",
        "low_threshold": "> 4000 for 5 minutes",
        "high_threshold": "> 8000 for 5 minutes",
        "description": "High disk write operations per second.",
    },
    {
        "name": "Disk read latency on Data Partition",
        "category": "Host",
        "low_threshold": "> 50ms for 5 minutes",
        "high_threshold": "> 100ms for 5 minutes",
        "description": "High disk read latency indicates storage performance issues.",
    },
    {
        "name": "Disk write latency on Data Partition",
        "category": "Host",
        "low_threshold": "> 50ms for 5 minutes",
        "high_threshold": "> 100ms for 5 minutes",
        "description": "High disk write latency indicates storage performance issues.",
    },
    {
        "name": "Disk space % used on Data Partition",
        "category": "Host",
        "low_threshold": "> 80% for 5 minutes",
        "high_threshold": "> 90% for 5 minutes",
        "description": "Disk space running low. May cause write failures.",
    },

    # System Resource Alerts
    {
        "name": "Swap Usage",
        "category": "Host",
        "low_threshold": "> 1GB for 5 minutes",
        "high_threshold": "> 2GB for 5 minutes",
        "description": "High swap usage indicates memory pressure.",
    },
    {
        "name": "System: CPU (User) %",
        "category": "Host",
        "low_threshold": "> 80% for 5 minutes",
        "high_threshold": "> 95% for 5 minutes",
        "description": "High CPU usage may indicate resource contention.",
    },
    {
        "name": "Page Faults",
        "category": "Host",
        "low_threshold": "> 100 for 5 minutes",
        "high_threshold": "> 500 for 5 minutes",
        "description": "High page faults indicate memory is being swapped to disk.",
    },

    # Queue Alerts
    {
        "name": "Queues: Readers",
        "category": "Host",
        "low_threshold": "> 10 for 5 minutes",
        "high_threshold": "> 50 for 5 minutes",
        "description": "Read operations waiting in queue. May indicate lock contention.",
    },
    {
        "name": "Queues: Writers",
        "category": "Host",
        "low_threshold": "> 10 for 5 minutes",
        "high_threshold": "> 50 for 5 minutes",
        "description": "Write operations waiting in queue. May indicate lock contention.",
    },

    # Ticket Alerts (WiredTiger)
    {
        "name": "Tickets Available: Reads",
        "category": "Host",
        "low_threshold": "< 50 for 5 minutes",
        "high_threshold": "< 20 for 5 minutes",
        "description": "Low read tickets available. May cause operation queueing.",
    },
    {
        "name": "Tickets Available: Writes",
        "category": "Host",
        "low_threshold": "< 50 for 5 minutes",
        "high_threshold": "< 20 for 5 minutes",
        "description": "Low write tickets available. May cause operation queueing.",
    },

    # Connection Alerts
    {
        "name": "Connections",
        "category": "Host",
        "low_threshold": "> 500 for 5 minutes",
        "high_threshold": "> 800 for 5 minutes",
        "description": "High number of connections. May indicate connection leak.",
    },

    # Assert Alerts
    {
        "name": "Asserts: Regular",
        "category": "Host",
        "low_threshold": "> 1 for 5 minutes",
        "high_threshold": "> 5 for 5 minutes",
        "description": "Assertion errors in MongoDB. May indicate bugs or issues.",
    },
    {
        "name": "Asserts: Warning",
        "category": "Host",
        "low_threshold": "> 5 for 5 minutes",
        "high_threshold": "> 10 for 5 minutes",
        "description": "Warning assertions. May indicate potential issues.",
    },

    # Ops Manager Backup Alerts
    {
        "name": "Backup Oplog Behind",
        "category": "Backup",
        "low_threshold": "Any occurrence",
        "high_threshold": None,
        "description": "Backup oplog is falling behind. Risk of backup gaps.",
    },
    {
        "name": "Backup Resync Required",
        "category": "Backup",
        "low_threshold": "Any occurrence",
        "high_threshold": None,
        "description": "Backup resync is required. Full resync may be needed.",
    },
    {
        "name": "Backup Agent Down",
        "category": "Backup",
        "low_threshold": "Any occurrence",
        "high_threshold": None,
        "description": "Backup agent is not running. Backups will not occur.",
    },

    # Agent Alerts (Ops Manager specific)
    {
        "name": "Monitoring Agent Down",
        "category": "Agent",
        "low_threshold": "Any occurrence",
        "high_threshold": None,
        "description": "Monitoring agent is down. No metrics will be collected.",
    },
    {
        "name": "Automation Agent Down",
        "category": "Agent",
        "low_threshold": "Any occurrence",
        "high_threshold": None,
        "description": "Automation agent is down. Cannot apply configuration changes.",
    },

    # Restarts
    {
        "name": "Restarts last hour",
        "category": "Host",
        "low_threshold": "> 1 for 5 minutes",
        "high_threshold": "> 2 for 5 minutes",
        "description": "MongoDB has restarted recently. May indicate crashes.",
    },
]


def create_excel_template(output_path: str = "opsmanager_alert_configurations.xlsx"):
    """Create the Excel template with alert configurations."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alert Configurations"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "Alert Name",
        "Alert Type/Category",
        "Low Priority Threshold",
        "High Priority Threshold",
        "Key Insights",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions["A"].width = 40  # Alert Name
    ws.column_dimensions["B"].width = 20  # Category
    ws.column_dimensions["C"].width = 25  # Low Priority
    ws.column_dimensions["D"].width = 25  # High Priority
    ws.column_dimensions["E"].width = 60  # Description

    # Add data
    for row, alert in enumerate(ALERT_CONFIGS, 2):
        ws.cell(row=row, column=1, value=alert["name"]).border = thin_border
        ws.cell(row=row, column=2, value=alert["category"]).border = thin_border
        ws.cell(row=row, column=3, value=alert["low_threshold"]).border = thin_border
        ws.cell(row=row, column=4, value=alert["high_threshold"]).border = thin_border
        ws.cell(row=row, column=5, value=alert["description"]).border = thin_border

        # Wrap text for description
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save
    wb.save(output_path)
    print(f"Created Excel template: {output_path}")
    print(f"Total alerts configured: {len(ALERT_CONFIGS)}")


if __name__ == "__main__":
    import sys
    from pathlib import Path

    script_dir = Path(__file__).parent
    output_path = script_dir / "opsmanager_alert_configurations.xlsx"

    if len(sys.argv) > 1:
        output_path = Path(sys.argv[1])

    create_excel_template(str(output_path))
